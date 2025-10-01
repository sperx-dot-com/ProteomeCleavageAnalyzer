import re
import pandas as pd
from Bio import SeqIO
from Bio.SeqUtils.ProtParam import ProteinAnalysis
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import pymol
from pymol import cmd


class SequenceAnalyzer:
    """Core functionality for analyzing FASTA sequences."""

    @staticmethod
    def find_occurrences_regex(sequence, regex_pattern):
        """Return a list of (start, match) tuples for regex_pattern in sequence."""
        regex = re.compile(regex_pattern, re.IGNORECASE)
        return [(m.start(), m.group()) for m in regex.finditer(sequence)]

    @staticmethod
    def scan_fasta_for_regex(fasta_path, regex_pattern, p1_pattern=None, at_cterm=False, close_cterm=False):
        """
        Scan a FASTA file for the regex pattern with optional P1' filtering and C-terminal filters.
        For each gene, prioritizes variants in this order:
        1. Variants without "isoform" in the name
        2. Variants with "isoform 1" specifically
        3. The first variant encountered
        Returns a list of hit dictionaries.
        """
        hits = []
        try:
            # First pass: organize records by gene ID with prioritization
            gene_to_record = {}
            gene_to_priority = {}  # 0: no isoform, 1: isoform 1, 2: other isoform
            
            for record in SeqIO.parse(fasta_path, "fasta"):
                seq_str = str(record.seq).upper()
                
                # Skip pseudogenes
                if "[pseudo=true]" in record.description or "*" in seq_str or "X" in seq_str:
                    continue
                    
                gene_match = re.search(r'\[gene\s*=\s*([^]]+)\]', record.description)
                if not gene_match:
                    continue
                    
                gene_id = gene_match.group(1).strip()
                protein_match = re.search(r'\[protein\s*=\s*([^]]+)\]', record.description)
                protein_name = protein_match.group(1).strip() if protein_match else "Unknown"
                
                # Determine priority
                if not protein_match or "isoform" not in protein_name.lower():
                    priority = 0  # Highest priority: no isoform
                elif "isoform 1" in protein_name.lower():
                    priority = 1  # Second priority: isoform 1
                elif "isoform alpha" in protein_name.lower():
                    priority = 2  # Second priority: isoform 1    
                elif "isoform X1" in protein_name.lower():
                    priority = 3  # Second priority: isoform 1
                else:
                    priority = 4  # Lowest priority: other isoforms
                
                # Store or update based on priority
                if gene_id not in gene_to_record or priority < gene_to_priority.get(gene_id, 2):
                    gene_to_record[gene_id] = record
                    gene_to_priority[gene_id] = priority
            
            # Second pass: process the selected records
            for gene_id, record in gene_to_record.items():
                seq_str = str(record.seq).upper()
                occurrences = SequenceAnalyzer.find_occurrences_regex(seq_str, regex_pattern.upper())
                
                if not occurrences:
                    continue
                    
                protein_match = re.search(r'\[protein\s*=\s*([^]]+)\]', record.description)
                protein_name = protein_match.group(1).strip() if protein_match else "Unknown"
                
                for pos, match_str in occurrences:
                    if at_cterm and (pos + len(match_str) != len(seq_str)):
                        continue
                    if close_cterm and (pos + len(match_str) < len(seq_str) - 10):
                        continue
                        
                    p1_residue = seq_str[pos + len(match_str): pos + len(match_str) + 1]
                    if p1_pattern and not re.fullmatch(p1_pattern, p1_residue, re.IGNORECASE):
                        continue
                        
                    hit = {
                        "gene_id": gene_id,
                        "protein": protein_name,
                        "weight": ProteinAnalysis(seq_str).molecular_weight(),
                        "protein_size": len(seq_str),  # will be displayed as "Length/aa"
                        "occurrences": [(pos, match_str)],
                        "sequence": [seq_str[pos: pos + len(match_str)]],
                        "p1prime": [p1_residue],
                        "match_snippets": [seq_str[max(0, pos - 10): pos + len(match_str) + 10]],
                        "full_seq": seq_str
                    }
                    hits.append(hit)
                    
            return hits
        except Exception as e:
            raise RuntimeError(f"Failed to process FASTA file: {e}")

    @staticmethod
    def load_go_annotations(go_file):
        """
        Parse a GAF file and group GO annotations by gene symbol and qualifier.
        Returns a dict: { gene_symbol: { qualifier: set([GO_ID, ...]), ... } }
        """
        gene_to_go = {}
        try:
            with open(go_file, 'r') as f:
                for line in f:
                    if line.startswith("!"):
                        continue
                    parts = line.strip().split("\t")
                    if len(parts) < 5:
                        continue
                    gene_symbol, qualifier, go_id = parts[2], parts[3], parts[4]
                    gene_to_go.setdefault(gene_symbol, {}).setdefault(qualifier, set()).add(go_id)
            return gene_to_go
        except Exception as e:
            raise RuntimeError(f"Failed to load GO annotations: {e}")

    @staticmethod
    def load_go_translations(obo_file):
        """
        Parse a GO OBO file and return a dict mapping GO IDs to term names.
        """
        go_translations = {}
        try:
            with open(obo_file, 'r') as f:
                current_id = None
                for line in f:
                    line = line.strip()
                    if line == "[Term]":
                        current_id = None
                    elif line.startswith("id: GO:"):
                        current_id = line.split("id: ")[1]
                    elif line.startswith("name: ") and current_id:
                        name = line.split("name: ")[1]
                        go_translations[current_id] = name
            return go_translations
        except Exception as e:
            raise RuntimeError(f"Failed to load GO translations: {e}")

    @staticmethod
    def annotate_hits(hits, go_annotations, go_translations):
        """
        For each hit, attach grouped GO annotations (with term names) using the provided dictionaries.
        """
        for hit in hits:
            gene_symbol = hit["gene_id"]
            if gene_symbol in go_annotations:
                grouped_go = {}
                for qualifier, go_ids in go_annotations[gene_symbol].items():
                    grouped_go[qualifier] = []
                    for go_id in go_ids:
                        translation = go_translations.get(go_id, "Unknown")
                        grouped_go[qualifier].append((go_id, translation))
                hit["go_terms"] = grouped_go
            else:
                hit["go_terms"] = {}
        return hits

    @staticmethod
    def mark_essential_genes(hits, excel_file):
        """
        Check if each hit's gene is present in the essential genes Excel file (assumes gene names in column 3).
        Adds an 'Essential' field to each hit.
        """
        try:
            df = pd.read_excel(excel_file, usecols=[2])
            essential_genes = set(df.iloc[:, 0].dropna().astype(str))
            for hit in hits:
                hit["Essential"] = "Yes" if hit["gene_id"] in essential_genes else "No"
            return hits
        except Exception as e:
            raise RuntimeError(f"Failed to load essential genes file: {e}")

    @staticmethod
    def get_reduced_hits(hits):
        """Return a reduced list of hits (first occurrence per gene)."""
        seen = set()
        reduced = []
        for hit in hits:
            if hit["gene_id"] not in seen:
                seen.add(hit["gene_id"])
                reduced.append(hit)
        return reduced

    @staticmethod
    def export_hits_to_excel(hits, output_file):
        """
        Export hits to an Excel file with formatted columns.
        The columns are: Gene ID, Protein, Weight/kDa, Length/aa, Essential, Position, 
        Matched Sequence, P1', Broader Sequence, GO Annotations, Full Sequence.
        """
        data = []
        for hit in hits:
            data.append({
                "Gene ID": hit['gene_id'],
                "Protein": hit['protein'],
                "Weight/kDa": hit['weight'] / 1000,
                "Length/aa": hit['protein_size'],
                "Essential": hit.get('Essential', "No"),
                "Position": hit['occurrences'][0][0]+1,
                "Matched Sequence": hit['sequence'][0],
                "P1'": hit['p1prime'][0],
                "Broader Sequence": hit['match_snippets'][0],
                "Gene Ontology": "; ".join([f"{term}" for _, terms in hit["go_terms"].items() 
                                          for go_id, term in terms]),
                "Full Sequence": hit['full_seq']
            })
        df = pd.DataFrame(data)
        wb = Workbook()
        ws = wb.active
        ws.title = "Gene Hits"
        
        
        # Find the index of the Gene ID column
        gene_id_col_idx = 1  # Assuming Gene ID is the first column
    
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply bold to header row
                if r_idx == 1:
                    cell.font = Font(bold=True)
                # Apply italic to Gene ID column (skip header row)
                elif c_idx == gene_id_col_idx and r_idx > 1:
                    cell.font = Font(italic=True)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        try:
            wb.save(output_file)
            return True, f"Exported {len(df)} hits to {output_file}"
        except Exception as e:
            raise RuntimeError(f"Export failed: {e}")
        
    @staticmethod
    def export_structure_data_to_excel(hits, file_path):
        """
        Export structure analysis data to Excel
        
        Parameters:
        hits (list): List of hit dictionaries with structure analysis data
        file_path (str): Path to save the Excel file
        
        Returns:
        tuple: (success, message)
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font
            
            # Prepare data for export
            data = []
            for hit in hits:
                # Insert uniprot url
                hit["uniprot_url"]="https://www.uniprot.org/uniprotkb/"+ str(hit["uniprot_id"]) + "/entry"

                # Format secondary structure without commas
                sec_struct = ""
                if hit.get("secondary_structure"):
                    # Sort by position
                    positions = sorted(hit["secondary_structure"].keys())
                    sec_struct = ''.join(hit["secondary_structure"][pos] for pos in positions)
                else:
                    sec_struct = "N/A"
                
                # Get GO terms text
                go_text = ""
                if hit.get("go_terms"):
                    parts = []
                    for qualifier, terms in hit["go_terms"].items():
                        for go_id, term in terms:
                            parts.append(f"{term}")
                    go_text = "; ".join(parts)
                
                data.append({
                    # Sequence data columns
                    "Gene ID": hit["gene_id"],
                    "Uniprot ID": hit["uniprot_id"],
                    "Protein": hit["protein"],
                    "Uniport Link": hit['uniprot_url'],
                    "Weight/kDa": f"{hit['weight']/1000:.1f}",
                    "Length/aa": hit["protein_size"],
                    "Position": hit["occurrences"][0][0]+1,
                    "Matched Sequence": hit["sequence"][0],
                    "P1' Residue": hit["p1prime"][0],
                    # "Context": hit["match_snippets"][0],
                    "Essential": hit.get("Essential", "N/A"),
                    
                    
                    # Structure data columns
                    "Total SASA": hit.get("total_sasa", "N/A"),
                    "Secondary Structure": sec_struct,
                    "Residues within 5Å": hit.get("residues_within_5.0A", "N/A"),
                    "Residues within 10Å": hit.get("residues_within_10.0A", "N/A"),
                    "Residues within 15Å": hit.get("residues_within_15.0A", "N/A"),
                    "GO Annotations": go_text

                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create and format Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Structure Analysis"
            
            # Find the index of the Gene ID column
            gene_id_col_idx = 1  # Assuming Gene ID is the first column
        
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Apply bold to header row
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                    # Apply italic to Gene ID column (skip header row)
                    elif c_idx == gene_id_col_idx and r_idx > 1:
                        cell.font = Font(italic=True)
                    
                    #attempt to set as hyperlinks
                    elif c_idx == gene_id_col_idx and r_idx > 1:
                        cell.font = Font(italic=True)
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
                
            wb.save(file_path)
            return True, f"Exported {len(df)} structure analysis records to {file_path}"
        except Exception as e:
            return False, f"Failed to export structure data: {str(e)}"
        
# Function to process arguments whether they're from command line or hardcoded
def process_sequence_analysis(args):
    try:
        print(f"Scanning {args.fasta} for pattern: {args.regex}")
        
        p1_regex = None
        if args.p1:
            p1_regex = f"[{args.p1.replace(',', '')}]"
        
        # Scan for regex matches
        hits = SequenceAnalyzer.scan_fasta_for_regex(
            args.fasta, args.regex, p1_regex, args.at_cterm, args.close_cterm
        )
        
        if not hits:
            print("No hits found.")
            return
            
        # Add GO annotations if files are provided
        if args.go and args.obo:
            print(f"Adding GO annotations from {args.go}")
            go_ann = SequenceAnalyzer.load_go_annotations(args.go)
            go_trans = SequenceAnalyzer.load_go_translations(args.obo)
            hits = SequenceAnalyzer.annotate_hits(hits, go_ann, go_trans)
        
        # Mark essential genes if file is provided
        if args.essential:
            print(f"Marking essential genes from {args.essential}")
            hits = SequenceAnalyzer.mark_essential_genes(hits, args.essential)
        
        # Reduce to one hit per gene
        hits = SequenceAnalyzer.get_reduced_hits(hits)
        
        # Export to Excel
        print(f"Found {len(hits)} unique gene hits")
        from structure_analyser import analyze_protein_region
        for hit in hits:
            #print(hit)
            length = len(hit['sequence'][0])
            start = hit['occurrences'][0][0]
            end = start + length
            structure_analysis = analyze_protein_region(hit["gene_id"],start, end )
            hit.update(structure_analysis)
            print(f"concat hit: {hit}")
            print("--------------------------------------")
        if args.output:
            SequenceAnalyzer.export_hits_to_excel(hits, args.output)
            print(f"Results exported to {args.output}")
        return hits
        
    except Exception as e:
        print(f"Error: {e}")
        return None


# Add a command-line interface
if __name__ == "__main__":
    import argparse
    import sys
    
    # Define test values for development/testing
    TEST_MODE = True  # Set to True to use hardcoded values, False to use command line args
    
    # Test values for development
    test_values = {
        "fasta": "genomes/ecoli_bl21de3.faa",
        "regex": "DEED",
        "p1": "G",
        "at_cterm": False,
        "close_cterm": False,
        "go": "go_files/ecocyc.gaf",
        "obo": "go_files/go.obo",
        "essential": "essential_ecoli.xlsx",
        "output": None
    }  
    parser = argparse.ArgumentParser(description="FASTA Sequence Analyzer")
    parser.add_argument("--fasta", help="Path to FASTA file")
    parser.add_argument("--regex", help="Regex pattern to search for")
    parser.add_argument("--p1", help="P1' pattern (comma separated)")
    parser.add_argument("--at-cterm", action="store_true", help="Only match at C-terminus")
    parser.add_argument("--close-cterm", action="store_true", help="Only match close to C-terminus (within 10 AA)")
    parser.add_argument("--go", help="Path to GO annotation file (GAF)")
    parser.add_argument("--obo", default="go_files/go.obo", help="Path to GO OBO file")
    parser.add_argument("--essential", help="Path to essential genes Excel file")
    parser.add_argument("--output", help="Output Excel file path")
    parser.add_argument("--use-test-values", action="store_true", help="Use hardcoded test values")
    
    args = parser.parse_args()
    
    # If --use-test-values is provided via command line, use test values
    if args.use_test_values or (TEST_MODE and len(sys.argv) == 1):
        # Create a namespace object with test values
        from types import SimpleNamespace
        args = SimpleNamespace(**test_values)
        print("Using test values for development")
    else:
        # Validate required args when not using test values
        if not args.fasta or not args.regex:
            parser.error("--fasta and --regex are required when not using test values")
    
    # Process the analysis with the provided arguments
    process_sequence_analysis(args)

class GOFilter:
    """Class to handle GO filtering business logic"""
    
    @staticmethod
    def filter_hits_by_expression(hits, expression):
        """Filter hits based on a GO expression with AND/OR/NOT logic"""
        filtered_hits = []
        
        # Normalize expression
        expression = expression.upper()
        
        # Special case for simple expressions without operators
        if "AND" not in expression and "OR" not in expression and "NOT" not in expression:
            # Single term search
            term = expression.strip()
            for hit in hits:
                if GOFilter.hit_matches_go_term(hit, term):
                    filtered_hits.append(hit)
            return filtered_hits
        
        # Handle complex expressions
        # First, split by OR
        or_parts = expression.split(" OR ")
        
        for hit in hits:
            # Check each OR clause - if any match, include the hit
            should_include = False
            
            for or_part in or_parts:
                or_part = or_part.strip()
                
                # Process AND parts within this OR clause
                if " AND " in or_part:
                    and_parts = or_part.split(" AND ")
                    and_result = True
                    
                    for and_part in and_parts:
                        and_part = and_part.strip()
                        
                        # Handle NOT operator
                        if and_part.startswith("NOT "):
                            term = and_part[4:].strip()
                            if GOFilter.hit_matches_go_term(hit, term):
                                and_result = False
                                break
                        else:
                            # Regular term
                            if not GOFilter.hit_matches_go_term(hit, and_part):
                                and_result = False
                                break
                    
                    if and_result:
                        should_include = True
                        break
                else:
                    # Simple OR term
                    # Check for NOT
                    if or_part.startswith("NOT "):
                        term = or_part[4:].strip()
                        if not GOFilter.hit_matches_go_term(hit, term):
                            should_include = True
                            break
                    else:
                        # Regular term
                        if GOFilter.hit_matches_go_term(hit, or_part):
                            should_include = True
                            break
            
            if should_include:
                filtered_hits.append(hit)
        
        return filtered_hits

    @staticmethod
    def hit_matches_go_term(hit, term):
        """Check if a hit matches a GO term or description"""
        if not hit.get("go_terms"):
            return False
        
        # Check if term is a GO ID
        if term.startswith("GO:"):
            # Look for exact GO ID match
            for qualifier, terms in hit["go_terms"].items():
                for go_id, description in terms:
                    if go_id == term:
                        return True
            return False
        else:
            # Look for term in any GO description
            term = term.lower()
            for qualifier, terms in hit["go_terms"].items():
                for go_id, description in terms:
                    if term in description.lower():
                        return True
            return False